from dataclasses import dataclass
import json
import os
from typing import Tuple, List, Optional, Dict
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill
import argparse
from pathlib import Path

MAX_VALUE = 3000
MIN_VALUE = 2000

@dataclass
class DataResult:
    disk_name: str = None
    rand_r_w_iops: Tuple[Optional[str], Optional[str]] = (None, None)
    seq_r_w_bw: Tuple[Optional[str], Optional[str]] = (None, None)
    randrw_iops: Optional[str] = None

    def __str__(self) -> str:
        rand_r = self.rand_r_w_iops[0] if self.rand_r_w_iops[0] else "?"
        rand_w = self.rand_r_w_iops[1] if self.rand_r_w_iops[1] else "?"
        seq_r = self.seq_r_w_bw[0] if self.seq_r_w_bw[0] else "?"
        seq_w = self.seq_r_w_bw[1] if self.seq_r_w_bw[1] else "?"
        randrw = self.randrw_iops if self.randrw_iops else "?"
        return (
            f"rand_r_w_iops: {rand_r}/{rand_w}\n"
            f"seq_r_w_bw: {seq_r}/{seq_w}\n"
            f"randrw_iops: {randrw}"
        )

def parse_json_file(file_path: str) -> Optional[Dict]:
    """Читает и парсит JSON-файл, возвращает данные или None при ошибке."""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            if not data.get("jobs"):
                print(f"Предупреждение: файл {file_path} не содержит данных 'jobs'.")
                return None
            return data
    except Exception as e:
        print(f"Ошибка при обработке файла {file_path}: {e}")
        return None

def process_reports(root_dir: str) -> Dict[str, Dict[str, DataResult]]:
    """
Обрабатывает JSON-отчеты FIO, группируя данные по узлам и дискам.
Возвращает словарь: {node_name: {disk_name: DataResult}}.
    """
    results = {}

    for dirpath, dirnames, filenames in os.walk(root_dir):
        parts = Path(dirpath).relative_to(root_dir).parts
        if len(parts) < 1:
            continue
        node_name = parts[0]
        disk_name = parts[1] if len(parts) > 1 else None

        if not disk_name:
            continue

        json_files = [f for f in filenames if f.endswith('.json')]
        if not json_files:
            continue

        if node_name not in results:
            results[node_name] = {}

        result = DataResult(disk_name=disk_name)
        read_iops = write_iops = seq_read_bw = seq_write_bw = None

        for filename in json_files:
            file_path = os.path.join(dirpath, filename)
            data = parse_json_file(file_path)
            if not data:
                continue

            job = data["jobs"][0]
            rw_type = job.get("job options", {}).get("rw", "")

            if "rr_4k" in filename:
                read_iops = f"{round(job['read']['iops'] / 1000, 1)}k"
            elif "rw_4k" in filename and rw_type == "randwrite":
                write_iops = f"{round(job['write']['iops'] / 1000, 1)}k"
            elif "wg_test" in filename and rw_type == "randrw":
                result.randrw_iops = f"{round(job['read']['iops'] / 1, 1)}"
            elif rw_type == "read":
                seq_read_bw = f"{round(job['read']['bw'] / 1024, 1)}"
            elif rw_type == "write":
                seq_write_bw = f"{round(job['write']['bw'] / 1024, 1)}"

        result.rand_r_w_iops = (read_iops, write_iops)
        result.seq_r_w_bw = (seq_read_bw, seq_write_bw)
        results[node_name][disk_name] = result

    return results

def calculateGradientColor(seqValue: float) -> str:
    """Вычисляет градиентный цвет между красным (#FF0000) и бледно-зеленым (#B5E6A2) через желтый."""
    # Пороговые значения
    min_val, max_val = MIN_VALUE, MAX_VALUE
    # Цвета: красный (#FF0000), желтый (#FFFF00), бледно-зеленый (#B5E6A2)
    colors = [
        (255, 0, 0),      # Красный при 5000
        (255, 255, 0),    # Желтый при 5750 (середина)
        (181, 230, 162)   # Бледно-зеленый при 6500
    ]

    # Нормализация значения в диапазоне [0, 1]
    t = (seqValue - min_val) / (max_val - min_val)
    t = max(0, min(1, t))  # Ограничим t в [0, 1]

    if t < 0.5:
        # Интерполяция между красным и желтым
        t = t * 2  # Масштабируем до [0, 1] для первого сегмента
        r = int(colors[0][0] + (colors[1][0] - colors[0][0]) * t)
        g = int(colors[0][1] + (colors[1][1] - colors[0][1]) * t)
        b = int(colors[0][2] + (colors[1][2] - colors[0][2]) * t)
    else:
        # Интерполяция между желтым и бледно-зеленым
        t = (t - 0.5) * 2  # Масштабируем до [0, 1] для второго сегмента
        r = int(colors[1][0] + (colors[2][0] - colors[1][0]) * t)
        g = int(colors[1][1] + (colors[2][1] - colors[1][1]) * t)
        b = int(colors[1][2] + (colors[2][2] - colors[1][2]) * t)

    # Преобразуем RGB в HEX
    return f"{r:02x}{g:02x}{b:02x}".upper()

def get_cell_color(seq_r_str: Optional[str]) -> str:
    """Определяет цвет ячейки на основе значения seq_r (последовательная скорость чтения)."""
    if not seq_r_str or seq_r_str == "?":
        return "FFFFFF"  # Белый для отсутствующих данных

    # Извлекаем числовое значение из строки, например "500.0 MB/s" -> 500.0
    try:
        seq_r = float(seq_r_str.split()[0])  # Берем число до "MB/s"
    except (ValueError, IndexError):
        return "FFFFFF"  # Белый при ошибке парсинга

    if seq_r >= MAX_VALUE:
        print(f"grather then 3000: {seq_r}")
        return "B5E6A2"  # Бледно-зеленый
    elif seq_r < MIN_VALUE:
        print(f"Less then 3000: {seq_r}")
        return "FF0000"  # Красный
    else:
        print(f"Calculate: {seq_r}")
        return calculateGradientColor(seq_r)

def write_to_excel(results: Dict[str, Dict[str, DataResult]], output_file: str) -> None:
    """Записывает результаты в Excel с фиксированной шириной столбцов 34 и раскраской ячеек по seq_r."""
    wb = Workbook()
    ws = wb.active
    ws.title = "FIO Results"

    # Собираем уникальные имена дисков
    all_disks = set()
    for node_data in results.values():
        all_disks.update(node_data.keys())
    all_disks = sorted(all_disks)

    # Заголовки таблицы
    ws.cell(row=1, column=1, value="Node Name")
    for col, disk_name in enumerate(all_disks, start=2):
        ws.cell(row=1, column=col, value=disk_name)

    # Заполнение данных и раскраска ячеек
    for row, (node_name, node_data) in enumerate(results.items(), start=2):
        ws.cell(row=row, column=1, value=node_name)
        for col, disk_name in enumerate(all_disks, start=2):
            result = node_data.get(disk_name)
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(wrap_text=True)  # Включаем перенос текста
            if result:
                cell.value = str(result)
                # Раскраска на основе seq_r
                color = get_cell_color(result.seq_r_w_bw[0])
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            else:
                cell.value = "?/?\n?/?\n?"
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Автонастройка ширины для первого столбца (Node Name)
    max_length = 0
    for cell in ws['A']:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    ws.column_dimensions['A'].width = max_length + 2

    # Фиксированная ширина 34 для столбцов с данными (дисками)
    for col in range(2, len(all_disks) + 2):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 34

    # Автонастройка высоты строк
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 60

    wb.save(output_file)
    print(f"Результаты сохранены в {output_file}")

def main():
    parser = argparse.ArgumentParser(description="Парсинг отчетов FIO и создание Excel-таблицы")
    parser.add_argument('--input-dir', default='fio_results', help='Директория с JSON-отчетами FIO')
    parser.add_argument('--output-file', default='fio_summary.xlsx', help='Имя выходного Excel-файла')
    args = parser.parse_args()

    root_dir = Path(args.input_dir)
    if not root_dir.exists():
        print(f"Ошибка: директория {root_dir} не существует.")
        return

    results = process_reports(root_dir)

    if results:
        for node_name, node_data in results.items():
            print(f"Node: {node_name}")
            for disk_name, result in node_data.items():
                print(f"  Disk: {disk_name}")
                print(result)
                print()
    else:
        print("Не найдено JSON-файлов для обработки.")

    write_to_excel(results, args.output_file)

if __name__ == '__main__':
    main()